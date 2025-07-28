from openpyxl.styles import numbers

from copy import copy
import base64
import json
import logging
import os
import time
from datetime import datetime

import requests
from docxtpl import DocxTemplate
from lxml import html
from openpyxl import Workbook
from openpyxl.styles import Side, Border, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
# from selenium.webdriver import ActionChains
# from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from sqlalchemy import func, and_, exists, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

from app import db, photos
from app.case_management.models import CaseProperty
from app.database.models import Property, Assessment, PropertyPhoto, Owner, CmaResult, CmaTask
from app.database.models.property import PropertyOriginal
from app.rules.adjustments import Adjustment
from app.settings.models import AssessmentDate, GlobalSetting
from app.utils.constants import County, PROPERTY_STYLE_MAP, BASEMENT_TYPE_MAP, LOCATION_STYLE_MAP, \
    PROPERTY_WATER_CATEGORY
from app.utils.number_converter import base_to_dec, mask_to_dec, dec_to_base
from config import EVIDENCE_DIR


class ReportService:

    def __init__(self, template_path):
        template_path = template_path
        self._docx = DocxTemplate(template_path)

    def gen_cma_report(self, subject, comps, output_dir=EVIDENCE_DIR):
        # prepare context for template rendering
        context = self.get_template_context(subject, comps)

        self._docx.render(context)
        self._docx.save(output_dir / f"cma_report_{subject.apn}.docx")
        # docx2pdf(output_dir / f"cma_report_{subject.apn}.docx")

    def get_comp_context(self, idx, comp):
        return {
            f'c{idx + 1}_sec': comp.section,
            f'c{idx + 1}_bl': comp.block,
            f'c{idx + 1}_lot': comp.lot,
            f'c{idx + 1}_apn': comp.apn,
            f'c{idx + 1}_district': str(comp.school_district) if comp.school_district else '-',
            f'c{idx + 1}_address': comp.address_line_1,
            f'c{idx + 1}_proximity': self.format_proximity(comp),
            f'c{idx + 1}_class': comp.property_class,
            f'c{idx + 1}_age': comp.age,
            f'c{idx + 1}_style': self.format_property_style(comp.county, comp.property_style),
            f'c{idx + 1}_condo_view': self.format_condo_view(comp),
            f'c{idx + 1}_sale_price': self.format_price(comp.last_sale_price),
            f'c{idx + 1}_sale_date': self.format_date(comp.last_sale_date),
            f'c{idx + 1}_beds': self.format_int(comp.bedrooms),

            # report subject inventories
            f'c{idx + 1}_gla': self.format_int(comp.gla_sqft),
            f'c{idx + 1}_lot_size': self.format_lot_size(comp.lot_size),
            f'c{idx + 1}_baths': self.format_int(comp.full_baths),
            f'c{idx + 1}_hbaths': self.format_int(comp.half_baths),
            f'c{idx + 1}_garages': self.format_int(comp.garages),
            f'c{idx + 1}_water': '-',
            f'c{idx + 1}_base': self.format_basement(comp.county, comp.basement_type),
            f'c{idx + 1}_location': self.format_location(comp.county, comp.location),
            f'c{idx + 1}_time': self.format_time_value(comp, key=Adjustment.TIME, is_subject=False),

            # adjustments
            f'c{idx + 1}_adj_gla': self.format_adjustment(comp, Adjustment.GLA),
            f'c{idx + 1}_adj_lot': self.format_adjustment(comp, Adjustment.LOT),
            f'c{idx + 1}_adj_baths': self.format_adjustment(comp, Adjustment.FULL_BATH),
            f'c{idx + 1}_adj_hbaths': self.format_adjustment(comp, Adjustment.HALF_BATH),
            f'c{idx + 1}_adj_garages': self.format_adjustment(comp, Adjustment.GARAGE),
            f'c{idx + 1}_adj_base': self.format_adjustment(comp, Adjustment.BASEMENT),
            f'c{idx + 1}_adj_time': self.format_time_adjustment_value(comp, key=Adjustment.TIME),
            f'c{idx + 1}_adj_beds': self.format_adjustment(comp, Adjustment.BEDROOMS),

            # summary adjustments
            f'c{idx + 1}_adj_delta': self.format_price(comp.adjustment_delta_value),
            f'c{idx + 1}_net_adj': self.format_price(comp.adjusted_market_value),
        }

    def get_subject_context(self, subject, comps):
        global_settings = db.session.query(GlobalSetting).filter_by(county=subject.county).first()
        return {
            's_report_title': global_settings.settings.get('pdf_header'),
            's_county': subject.property_county.name.upper(),
            's_report_date': self.format_date(str(datetime.now().date())),
            's_sec': subject.section,
            's_bl': subject.block,
            's_lot': subject.lot,
            's_beds': self.format_int(subject.bedrooms),
            's_apn': subject.apn,
            's_district': str(subject.school_district) if subject.school_district else '-',
            's_address': subject.address_line_1,
            's_proximity': self.format_proximity(subject),
            's_class': subject.property_class,
            's_age': subject.age,
            's_style': self.format_property_style(subject.county, subject.property_style),
            's_condo_view': self.format_condo_view(subject),
            's_sale_price': self.format_price(subject.last_sale_price),
            's_sale_date': self.format_date(subject.last_sale_date),
            # 's_picture': ReportService.format_picture(subject),


            # report subject inventories
            's_gla': self.format_int(subject.gla_sqft),
            's_lot_size': self.format_lot_size(subject.lot_size),
            's_baths': self.format_int(subject.full_baths),
            's_hbaths': self.format_int(subject.half_baths),
            's_garages': self.format_int(subject.garages),
            's_water': '-',
            's_base': self.format_basement(subject.county, subject.basement_type),
            's_location': self.format_location(subject.county, subject.location),
            's_time': self.format_time_value(comps[0], key=Adjustment.TIME, is_subject=True),
        }

    def get_template_context(self, subject, comps):
        context = self.get_subject_context(subject, comps)

        for i in range(len(comps)):
            comp_context = self.get_comp_context(i, comps[i])
            context.update(comp_context)

        return context

    @classmethod
    def format_adjustment(cls, comp, key):
        if not hasattr(comp, 'adjustments'):
            return ''

        for adj in comp.adjustments:
            if adj['key'] == key:
                if not adj['adjusted']:
                    return ''
                return cls.format_price(adj['value'])

        return ''

    @classmethod
    def format_date(cls, value):
        if value is None:
            return '-'
        formatted = datetime.strptime(str(value), '%Y-%m-%d').strftime('%m/%d/%Y')
        return formatted

    @classmethod
    def format_picture(cls, prop):
        # This seams to clear the content in my cell
        # tables[0].rows[0].cells[0]._element.clear_content()

        # Then when the image is inserted to the cell it is not placed one linefeed down.
        # img = tables[0].rows[0].cells[0].add_paragraph().add_run().add_picture('Image.png', width=Inches(0.4))
        pass

    @classmethod
    def format_proximity(cls, prop):
        formatted = '-'
        if hasattr(prop, 'proximity'):
            if prop.proximity:
                formatted = '{:0,.2f} Miles'.format(prop.proximity)
        return formatted

    @classmethod
    def format_price(cls, value):
        if value is None:
            return '-'
        formatted = '${:0,.0f}'.format(value).replace('$-', '-$')
        return formatted

    @classmethod
    def format_condo_view(cls, prop):
        formatted = '-'
        if hasattr(prop, 'condo_view_influence') and hasattr(prop, 'condo_code_description'):
            if prop.condo_view_influence and prop.condo_code_description:
                formatted = f'({prop.condo_view_influence}) {prop.condo_code_description}'
        return formatted

    @classmethod
    def format_property_style(cls, county, style_code):
        if not county or not style_code:
            return '-'
        try:
            code = int(style_code)
            formatted = f'{PROPERTY_STYLE_MAP.get(county).get(code)}'
        except Exception as e:
            print(e)
            return '-'

        return formatted

    @classmethod
    def format_basement(cls, county, code):
        if not county or not code:
            return '-'
        try:
            code = int(code)
            formatted = f'{BASEMENT_TYPE_MAP.get(county).get(code).upper()}'
        except Exception as e:
            print(e)
            return '-'

        return formatted

    @classmethod
    def format_time_value(cls, prop, key, is_subject=False):
        if not hasattr(prop, 'adjustments'):
            return ''

        for adj in prop.adjustments:
            if adj['key'] == key:
                if not adj['adjusted']:
                    return ''

                if is_subject:
                    return cls.format_date(adj['subject_value'])
                else:
                    return cls.format_date(adj['comp_value'])

        return ''

    @classmethod
    def format_time_adjustment_value(cls, prop, key):
        if not hasattr(prop, 'adjustments'):
            return ''

        for adj in prop.adjustments:
            if adj['key'] == key:
                if not adj['adjusted']:
                    return ''

                return cls.format_price(adj['value'])

        return ''

    @classmethod
    def format_location(cls, county, code):
        if not county or not code:
            return '-'
        try:
            formatted = f'{LOCATION_STYLE_MAP.get(county).get(str(code)).upper()}'
        except Exception as e:
            print(e)
            if code:
                return str(code)
            return '-'

        return formatted

    @classmethod
    def format_int(cls, value):
        if value is None:
            return '-'

        formatted = '{:0,.0f}'.format(value)
        return formatted

    @classmethod
    def format_lot_size(cls, value):
        if value is None:
            return '-'
        formatted = '{:0,.3f}'.format(value)
        return formatted

    @classmethod
    def format_waterfront(cls, county, category):
        if category is None or county is None:
            return '-'

        formatted = f'({category}) {PROPERTY_WATER_CATEGORY.get(county).get(category)}'
        return formatted


class PropertyService:
    # add inventories, what can be modified
    overridden_inventories = ['gla_sqft', 'lot_size']

    @classmethod
    def get_last_mass_cma_result(cls, property_id):

        query = (
            db.session.query(CmaResult).join(
                CmaTask, isouter=True).filter(
                CmaResult.property_id == property_id).order_by(
                CmaTask.task_ts.desc())
            .limit(1)
        )

        cma_result = query.first()
        return cma_result

    @classmethod
    def get_unique_assessment_stage(self):
        max_stage = db.session.execute('''SELECT MAX(assessment_stage) FROM property;''')
        return max_stage + 1

    @classmethod
    def mark_mass_cma_stage(cls, properties, assessment_stage):

        if not properties:
            return None

        if len(properties) > 1:
            property_ids = f"{tuple(properties)}"
        else:
            # the only one value in a list, add () manually
            property_ids = f"({properties[0]})"

        query = f'''update property set assessment_stage = {assessment_stage} where id in {property_ids};'''
        print(query)
        db.session.execute(query)

    @classmethod
    def get_property_query(cls, **kwargs):
        """
        Get 'property' query object
        :param kwargs: Query filter parameters
        """
        # pop mass_cma param if in kwargs, default 'False'
        mass_cma = kwargs.pop('mass_cma', False)

        if mass_cma:
            # specific Mass CMA query configuration
            query = db.session.query(Property).filter_by(**kwargs)
        else:
            # specific Single CMA query configuration
            query = db.session.query(Property).options(
                joinedload('owners'),
                joinedload('photos')
            ).filter_by(**kwargs)

        return query

    @classmethod
    def get_filter_query(cls, **kwargs):
        """
        Get 'property' query object
        Does minimum selects/joining to lighten the filtering process
        :param kwargs: Query filter parameters
        """
        query = db.session.query(Property.id, Property.apn, Property.latitude, Property.longitude).filter_by(**kwargs)

        return query

    @classmethod
    def override_property_from_original(cls, prop):
        """
        Override property attributes from the original property data
        """
        # get original property data
        original_prop = db.session.query(PropertyOriginal).filter(PropertyOriginal.id == prop.id).first()

        # override target property with original data
        for key, value in original_prop.__dict__.items():
            if not key.startswith('_') and not callable(getattr(original_prop, key)):
                setattr(prop, key, value)
        db.session.commit()
        return prop

    @classmethod
    def original_property_exists(cls, property_id):
        """
        Check whether original property exist in 'property_original' table
        """
        return db.session.query(exists().where(PropertyOriginal.id == property_id)).scalar()

    @classmethod
    def _get_property_mappings(cls, prop):
        """
        Return property attributes key-value mappings
        """
        keys = db.inspect(Property).columns.keys()
        return {key: getattr(prop, key) for key in keys}

    @classmethod
    def backup_original_property(cls, property_id):
        """
        Backup original property data
        :param property_id: The property id to backup
        """
        # check if original property data exists in 'property_original' table
        if not cls.original_property_exists(property_id):
            props = Property.query.filter(Property.id == property_id)
            db.session.bulk_insert_mappings(PropertyOriginal, (cls._get_property_mappings(prop) for prop in props))
            db.session.commit()
            return True
        return False

    @classmethod
    def generate_code(cls, property_id: int):
        """
        Generate a code from property_id
        For each value add a mask '1A2B3C' value to avoid encoded 000001 like numbers
        """
        mask_value = mask_to_dec()

        value = int(property_id) + mask_value
        return dec_to_base(value)

    @classmethod
    def read_code(cls, code: str):
        # convert code to decimal numeric system
        try:
            code_number = base_to_dec(code)

            # sub a mask back
            property_id = code_number - mask_to_dec()
            if property_id and property_id > 0:
                return property_id
        except Exception as e:
            print(f'{e}: {code}')
        return None

    @classmethod
    def get_property(cls, property_id, original: bool = False, mass_cma: bool = False):
        """
        Get property
        :param property_id: The property id
        :param original: Whether get original property data
        :param mass_cma: Whether get specific MassCMA configured property
        """
        prop = cls.get_property_query(id=property_id, mass_cma=mass_cma).first()

        # get original property data
        if original and cls.original_property_exists(property_id):
            # override last updated property with original data
            return cls.override_property_from_original(prop)

        return prop

    @classmethod
    def get_property_last_assessment(cls, property_id, tax_year=None):
        """
        Get property last assessment
        """
        # join tables
        query = (
            db.session.query(Assessment).filter(Assessment.property_id == property_id).join(
                AssessmentDate,
                Assessment.assessment_id == AssessmentDate.id
            )
        )

        # filter by tax_year optionally
        if tax_year:
            query = query.filter(AssessmentDate.tax_year == tax_year)

        query = query.order_by(
            AssessmentDate.valuation_date.desc()
        ).limit(1)

        # last_assessment = (
        #     db.session.query(Assessment).filter(Assessment.property_id == property_id).join(
        #         AssessmentDate,
        #         Assessment.assessment_id == AssessmentDate.id
        #     ).order_by(
        #         AssessmentDate.valuation_date.desc()
        #     ).limit(1).correlate(Property)).first()

        return query.first()

    @classmethod
    def get_property_assessment(cls, property_id, assessment_date_id=None, tax_year=None):
        """
        Get property assessment
        """
        # get last property assessment
        if assessment_date_id is None:
            assessment = cls.get_property_last_assessment(property_id=property_id, tax_year=tax_year)

        # get property assessment by assessment date id
        else:
            assessment = db.session.query(Assessment).filter(
                Assessment.property_id == property_id,
                Assessment.assessment_id == assessment_date_id
            ).first()

        return assessment

    @classmethod
    def search_by_address(cls, query, address_string, address_attribute):
        """Get query object to search by address"""
        address_parts = address_string.strip().lower().replace(',', '').split()
        word_filters = []
        for part in address_parts:

            # if address part is numeric, then match it exactly
            if part.isnumeric():
                # query = query.filter(Property.address.match(part))
                if part[-1] == '1':
                    or_part = 'st'
                elif part[-1] == '2':
                    or_part = 'nd'
                elif part[-1] == '3':
                    or_part = 'rd'
                else:
                    or_part = 'th'
                query = query.filter(or_(
                    func.to_tsvector('english',
                                     address_attribute).match(part,
                                                              postgresql_regconfig='english'),
                    func.to_tsvector('english',
                                     address_attribute).match(part + or_part,
                                                              postgresql_regconfig='english')
                ))

            # if word, then check if address contains substring
            else:
                word_filters.append(func.lower(address_attribute).ilike(
                    f"%{part.strip()}%",
                    escape=','))

        if word_filters:
            # apply 'word_filters'
            query = query.filter(
                and_(
                    *word_filters
                )
            )

        # if word filter return 0 count do not apply it
        # for f in word_filters:
        #     if query.filter(f).count():
        #         query = query.filter(f)

        return query

    @classmethod
    def _search_by_block(cls, query, block: str):
        return query.filter(func.lower(Property.block) == block.lower())

    @classmethod
    def _search_by_lot(cls, query, lot: str):
        return query.filter(func.lower(Property.lot) == lot.lower())

    @classmethod
    def _search_by_section(cls, query, section: str):
        return query.filter(func.lower(Property.section) == section.lower())

    @classmethod
    def _search_by_district(cls, query, district: str):
        return query.filter(func.lower(Property.district) == district.lower())

    @classmethod
    def _search_by_apn(cls, query, apn: str):
        """
        Get query object to search by apn
        """

        apn_parts = apn.split()
        if len(apn_parts) == 1:
            return query.filter(Property.apn == apn)

        apn_filters = [
            func.lower(Property.apn).ilike(f"%{part.strip()}%") for part in apn_parts
        ]

        query = query.filter(
            and_(
                *apn_filters
            )
        )

        return query

    @classmethod
    def search_properties(cls, args, model=Property):
        """
        Search properties by filter criteria
        :param args: The search criteria arguments
        :param model: The property model
        """
        # pop limit from args, since its not a database argument
        limit = args.pop('limit', None)

        # get search property query
        query = cls.search_properties_query(args, model)

        # limit result set
        if limit:
            query = query.limit(limit)

        # fetch objects
        objects = query.all()
        return objects

    @classmethod
    def search_properties_query(cls, args, model=Property):

        # accept search only by 'id', 'zip', 'state', 'section', 'block', 'lot', 'address', 'county', 'apn
        filters = {
            k: v for k, v in args.items() if v is not None and k in [
                'id', 'apn', 'zip', 'state', 'section', 'block', 'lot', 'district',
                'address', 'limit', 'county', 'city', 'unit',
                'address_line_1', 'address_line_2'
            ]
        }

        # prepare string filters: lower(), strip()
        for k, v in filters.items():
            if isinstance(v, str):
                filters[k] = v.strip()
            else:
                filters[k] = v

        ignore_address_tail = args.get('ignore_address_tail')

        # pop address
        address = filters.pop('address', None) if not ignore_address_tail else None

        # pop address line 1
        address_line_1 = filters.pop('address_line_1', None) \
            if not ignore_address_tail else filters.pop('address', None)

        # pop address line 2
        address_line_2 = filters.pop('address_line_2', None) \
            if not ignore_address_tail else filters.pop('unit', None)

        # pop apn
        apn = filters.pop('apn', None)

        # pop county
        county = filters.pop('county', None)

        # pop block
        block = filters.pop('block', None)

        # pop lot
        lot = filters.pop('lot', None)

        # pop section
        section = filters.pop('section', None)

        district = filters.pop('district', None)

        # filter query by args except 'address'
        query = db.session.query(model).filter_by(**filters)

        # filter all 'gis' records
        query = query.filter(model.origin.is_(None))

        # search by county
        if county:
            county_code = County.get_code(county)
            query = query.filter_by(county=county_code)

        # search by address
        if address:
            query = cls.search_by_address(query, address, address_attribute=model.address)

        # search by address_line_1
        if address_line_1:
            query = cls.search_by_address(query, address_line_1, address_attribute=model.address_line_1)

        # search by address_line_2
        if address_line_2:
            query = cls.search_by_address(query, address_line_2, address_attribute=model.address_line_2)

        # search by apn
        if apn:
            query = cls._search_by_apn(query, apn)

        if block:
            query = cls._search_by_block(query, block)

        if lot:
            query = cls._search_by_lot(query, lot)

        if section:
            query = cls._search_by_section(query, section)

        if district:
            query = cls._search_by_district(query, district)

        return query

    @classmethod
    def get_property_owners(cls, property_id):
        owners = db.session.query(Owner).filter(Owner.property_id == property_id).filter(
            Owner.data_source == 'assessment'
        ).order_by(Owner.created_on.desc()).all()
        return owners

    @classmethod
    def get_last_property_owner(cls, property_id):
        owners = cls.get_property_owners(property_id)
        if owners:
            return owners[0]
        return None

    @classmethod
    def get_last_owner_full_name(cls, property_id):
        last_owner = cls.get_last_property_owner(property_id)
        owner = ''
        if last_owner:
            owner1 = ' '.join([last_owner.first_name or '',
                               last_owner.last_name or '']).strip()
            owner2 = ' '.join([last_owner.second_owner_first_name or '',
                               last_owner.second_owner_last_name or '']).strip()
            owner = ','.join([owner1, owner2])

        return owner

    @classmethod
    def get_owners_full_names(cls, property_id):
        """
        Get first owner found
        """
        owners = cls.get_property_owners(property_id)

        _full_names = []
        for o in owners:
            f_names = o.get_full_names()
            for full_name in f_names:
                if full_name and full_name not in _full_names:
                    _full_names.append(full_name)

        return _full_names

    @classmethod
    def get_matched_owner(cls, property_id, first_name, last_name):
        owners = cls.get_property_owners(property_id)

        for o in owners:
            if str(o.first_full_name):
                if first_name.lower() in str(o.first_full_name).lower() and \
                        last_name.lower() in str(o.first_full_name).lower():
                    return o
            if str(o.second_full_name):
                if first_name.lower() in str(o.second_full_name).lower() and \
                        last_name.lower() in str(o.second_full_name).lower():
                    return o
        return None


class PhotoService:

    @classmethod
    def get_best_photo(cls, property_id):
        """
        Get best property photo
        :param property_id: The property id
        """
        return PropertyPhoto.query.filter_by(property_id=property_id, is_best=True).one()

    @classmethod
    def get_photo_name(cls, file):
        """
        Get photo name
        """
        now = datetime.now()
        time_stamp = now.strftime("%H%M%S")
        prefix, _, extension = file.filename.rpartition('.')

        return f'{prefix}-upload-{time_stamp}.{extension}'

    @classmethod
    def make_photo_url(cls, prop, photo):
        """
        Make property photo url
        """
        return photos.url(f'{prop.county}/photos/{photo.name}')

    @classmethod
    def make_all_photos_urls(cls, prop):
        """
        Make property all photos urls
        """
        if hasattr(prop, 'photos') and prop.photos:
            for photo in prop.photos:
                photo.url = PhotoService.make_photo_url(prop, photo)
        return prop

    @classmethod
    def get_rank_photos(cls, property_id):
        """
        Get property rank photos.
        """
        rank_photos = db.session.query(
            PropertyPhoto,
            func.rank().over(
                order_by=PropertyPhoto.created_at.desc(),
                partition_by=PropertyPhoto.property_id
            ).label('rank')
        ).filter_by(property_id=property_id).all()

        all_photos = []
        for rank_photo in rank_photos:
            photo = rank_photo[0]

            # set best photo rank -1
            photo.rank = -1 if photo.is_best else rank_photo.rank

            all_photos.append(photo)
        return all_photos


class ScraperService:
    def create_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")

        driver = webdriver.Chrome(options=chrome_options)

        return driver

    def broward_evidence(self, apn, gen_pdf=True, gen_html=True):
        driver = self.create_driver()
        driver.get(f'https://bcpa.net/RecInfo.asp?URL_Folio={apn}')
        # driver.get("http://web.bcpa.net/BcpaClient/#/Record-Search")

        # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "txtField")))
        #
        # search_input = driver.find_element_by_id('txtField')
        # search_input.send_keys(apn)
        # search_input.send_keys(Keys.ENTER)
        #
        # try:
        #     WebDriverWait(driver, 7).until(EC.element_to_be_clickable((
        #         By.XPATH,
        #         '//*[@id="hideRecInfoTab"]/div[1]/div/div[12]'))).click()
        # except TimeoutException:
        #     print(f'Property {apn} not found')
        #     return
        #
        # driver.implicitly_wait(5)
        #
        # print(f'opened windows {driver.window_handles}')
        # driver.switch_to.window(driver.window_handles[1])
        #
        # time.sleep(3)

        pdf_path = None
        if gen_pdf:
            path = self.gen_storage_path(apn)
            pdf_path = self._gen_pdf(driver, apn, path)

        if gen_html:
            path = self.gen_storage_path(apn)
            self._gen_html(driver, apn, path)

        time.sleep(3)
        driver.quit()

        return pdf_path

    def miamidade_evidence(self, apn, gen_pdf=True, gen_html=True):
        prop = Property.query.filter_by(apn=apn).first()
        if not prop:
            print(f'property with apn {apn} does not exists')
            exit(1)

        driver = self.create_driver()
        driver.get('https://www.miamidade.gov/Apps/PA/propertysearch/#/')

        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "t-folio"))).click()

        driver.implicitly_wait(5)

        search_input = driver.find_element_by_xpath(
            '/html/body/div/div[2]/div[3]/div[1]/div[2]/div[2]/div/div[4]/div/input')
        search_input.send_keys(apn)
        search_input.send_keys(Keys.ENTER)

        driver.implicitly_wait(5)

        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "print-modal"))).click()

        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "/html/body/div[1]/div[2]/div[3]/div[4]/div/div/div[2]/div[1]/a"))).click()

        driver.implicitly_wait(5)
        driver.switch_to.window(driver.window_handles[1])

        print('wait for image load')
        time.sleep(5)

        pdf_path = None
        if gen_pdf:
            path = self.gen_storage_path(apn)
            pdf_path = self._gen_pdf(driver, apn, path)

        if gen_html:
            path = self.gen_storage_path(apn)
            self._gen_html(driver, apn, path)

        time.sleep(2)
        driver.quit()

        return pdf_path

    def gen_storage_path(self, apn):
        storage_path = EVIDENCE_DIR / apn
        os.makedirs(storage_path, exist_ok=True)

        return storage_path

    def _gen_pdf(self, driver, apn, path):
        # store as pdf
        # See https://chromedevtools.github.io/devtools-protocol/tot/Page/#method-printToPDF for all options
        save_as_pdf(driver, f'{path}/{apn}.pdf', {'landscape': False, 'displayHeaderFooter': False})

        return f'{path}/{apn}.pdf'

    def _gen_html(self, driver, apn, path):
        # store as html
        content = driver.page_source
        # write the page content
        with open(f'{path}/{apn}.html', 'w') as fp:
            fp.write(content)

        # download the referenced files to the same path as in the html
        sess = requests.Session()
        base = 'https://web.bcpa.net/BcpaClient/'
        sess.get(base)  # sets cookies

        # parse html
        h = html.fromstring(content)
        # get css/js files loaded in the head
        for hr in h.xpath('head//@href'):
            if not hr or hr.startswith('http'):
                continue
            local_path = f'{path}/' + hr
            hr = base + hr
            res = sess.get(hr)
            if not os.path.exists(os.path.dirname(local_path)):
                os.makedirs(os.path.dirname(local_path))
            with open(local_path, 'wb') as fp:
                fp.write(res.content)

        # get image/js files from the body.  skip anything loaded from outside sources
        for src in h.xpath('//@src'):
            if not src or src.startswith('http'):
                continue
            local_path = f'{path}/' + src
            print(local_path)
            src = base + src
            res = sess.get(src)
            if not os.path.exists(os.path.dirname(local_path)):
                os.makedirs(os.path.dirname(local_path))
            with open(local_path, 'wb') as fp:
                fp.write(res.content)


class HearingService:
    @classmethod
    def parse_petition_row(cls, row):
        """
        Parse petition row data of the uploading file
        """
        # parse date/time object
        date_time = row['hearing_date']
        hearing_date = None
        hearing_time = None

        if date_time:
            # parse hearing date
            hearing_date = date_time.date().strftime('%m/%d/%Y')

            # parse hearing time
            hearing_time = date_time.strftime('%I:%M%p').lstrip('0')

        # parse apn
        apn = row['apn'].replace('-', '')

        # parse petition number
        petition_number = row['petition_number']

        # board room
        board_room = row['board_room']

        # extend row with new parsed columns
        row['apn'] = apn or None
        row['hearing_date'] = hearing_date
        row['hearing_time'] = hearing_time
        row['petition_number'] = petition_number or None
        row['board_room'] = board_room or None

        return row

    @staticmethod
    def any_valid(*args):
        for arg in args:
            if arg is not None:
                return True
        return False

    @classmethod
    def persist_petition_data(cls, df):
        for index, row in df.iterrows():
            apn = row['apn']
            case = CaseProperty.get_by(apn=apn)
            try:
                # update if exists only
                hearing_date = row.get('hearing_date')
                hearing_time = row.get('hearing_time')
                petition_number = row.get('petition_number')
                board_room = row.get('board_room')

                # if any is present, then override
                if case and cls.any_valid(hearing_date, hearing_time, petition_number, board_room):
                    case.hearing_date = hearing_date
                    case.hearing_time = hearing_time
                    case.petition_number = petition_number
                    case.board_room = board_room
                    case.save()

                    logging.info(f'updated hearing data for the case with apn={apn}')
                else:
                    logging.warning(f'Case object not found for apn={apn}')
            except IntegrityError as e:
                db.session.rollback()
                logging.error(e.orig.args)
        return None


class PetitionsReportService(object):

    # accounting excel number format
    # accounting_number_format = u'_($* #,##0.00_);[Black]_($* (#,##0.00);_($* -_0_0_);_(@'
    accounting_number_format = u'_($* #,##0.00_);[Black]_($* (#,##0.00);_($* #,##0.00_);_(@'

    # two colorized rows background fill
    background_fill = PatternFill("solid", fgColor="D9D9D9")

    # font for the 'FOLIO' specific column
    tahoma_font = Font(name='Tahoma', size=9)

    # default cell font
    cell_font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                     strike=False, color='FF000000')

    # bold cell font
    cell_font_bold = copy(cell_font)
    cell_font_bold.bold = True

    # underlined cell font
    cell_font_underlined = copy(cell_font)
    cell_font_underlined.underline = 'single'
    cell_font_underlined.color = '0563C1'

    # cell alignment
    cell_alignment = Alignment(horizontal='left', vertical='top', text_rotation=0, wrap_text=False,
                               shrink_to_fit=False, indent=0)

    # centered cell alignment
    cell_alignment_centered = copy(cell_alignment)
    cell_alignment_centered.horizontal = 'center'

    # header font
    header_font = Font(color='FFFFFF', size=11, name='Calibri', bold=True)

    # no border
    side = Side(border_style=None)
    no_border = Border(left=side, right=side, top=side, bottom=side)

    # header pattern fill
    header_fill = PatternFill("solid", fgColor="000000")

    # header alignment
    header_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False,
                                 shrink_to_fit=False, indent=0)

    def __init__(self):
        self._workbook = Workbook()
        self._ws = self._workbook.active
        self._rows = []

    @property
    def workbook(self):
        return self._workbook

    def setup_summary_worksheet(self, title='Summary'):

        rows_count = len(self._rows) + 2

        # Select active Worksheet
        self._ws.title = title

        # create and style the header
        self._set_header_styling()

        # currency columns styling
        self._set_column_styling(max_row=rows_count, min_col=7, max_col=12)
        self._set_column_styling(max_row=rows_count, min_col=14, max_col=14)
        self._set_column_styling(max_row=rows_count, min_col=16, max_col=16)
        self._set_column_styling(max_row=rows_count, min_col=18, max_col=18)
        self._set_column_styling(max_row=rows_count, min_col=20, max_col=20)

        # percent columns styling
        self._set_column_styling(max_row=rows_count, min_col=13, max_col=13, style='Percent',
                                 number_format=numbers.FORMAT_PERCENTAGE, alignment=self.cell_alignment_centered)
        self._set_column_styling(max_row=rows_count, min_col=15, max_col=15, style='Percent',
                                 number_format=numbers.FORMAT_PERCENTAGE, alignment=self.cell_alignment_centered)
        self._set_column_styling(max_row=rows_count, min_col=17, max_col=17, style='Percent',
                                 number_format=numbers.FORMAT_PERCENTAGE, alignment=self.cell_alignment_centered)
        self._set_column_styling(max_row=rows_count, min_col=19, max_col=19, style='Percent',
                                 number_format=numbers.FORMAT_PERCENTAGE, alignment=self.cell_alignment_centered)
        self._set_column_styling(max_row=rows_count, min_col=21, max_col=21, style='Percent',
                                 number_format=numbers.FORMAT_PERCENTAGE, alignment=self.cell_alignment_centered)

        #  Settle | Court  column styling
        self._set_column_styling(max_row=rows_count, min_col=22, max_col=22, alignment=self.cell_alignment_centered)

        # counter column styling
        right_border = copy(self.no_border)
        right_border.right = Side(border_style='thin')
        self._set_column_styling(max_row=rows_count, min_col=1, max_col=1, font=self.cell_font_bold, style='Normal',
                                 number_format=numbers.FORMAT_GENERAL, border=right_border)

        # 'FOLIO' and 'Petition Num' columns styling
        self._set_column_styling(max_row=rows_count, min_col=5, max_col=6, font=self.tahoma_font, style='Normal',
                                 number_format=numbers.FORMAT_GENERAL)

        # hyperlinks columns styling
        self._set_column_styling(max_row=rows_count, min_col=2, max_col=4, font=self.cell_font_underlined,
                                 alignment=self.cell_alignment_centered, style='Hyperlink',
                                 number_format=numbers.FORMAT_GENERAL)

        # style two colored rows
        self._set_background_styling(row_count=rows_count)

        # set columns dimensions
        self._set_columns_dimensions()

        self.set_outline_border(f'A3:V{rows_count}')

        # set zoom level
        self._ws.sheet_view.zoomScale = 85

        # self._workbook.save(output_path)
        return self._workbook

    def _insert_hyperlink_cell(self, row, col, value, hyperlink):
        if hyperlink and value:
            self._ws.cell(row=row, column=col).value = value
            self._ws.cell(row=row, column=col).hyperlink = hyperlink

    def _insert_row(self, position, data):
        """
        Insert row data
        """
        # insert counter
        self._ws.cell(row=position, column=1).value = position - 2

        self._insert_hyperlink_cell(row=position, col=2, value='Link to PDF', hyperlink=data[0])
        self._insert_hyperlink_cell(row=position, col=3, value='Link to PDF', hyperlink=data[1])
        self._insert_hyperlink_cell(row=position, col=4, value='Link to Single CMA', hyperlink=data[2])

        # insert values for the columns: 'FOLIO': 'Redux Proposals'
        for i in range(3, len(data)):
            self._ws.cell(row=position, column=i+2).value = data[i]

        # Insert formulas, based on inserted data

        # 'Redux Discount' formulas
        self._ws.cell(row=position, column=12).value = f'=$G{position}-$J{position}'
        self._ws.cell(row=position, column=13).value = f'=($G{position}-$J{position})/$G{position}'

        # 'County Discount' formulas
        self._ws.cell(row=position, column=14).value = f'=$G{position}-$K{position}'
        self._ws.cell(row=position, column=15).value = f'=($G{position}-$K{position})/$G{position}'

        # 'Difference All 1-4' formulas
        self._ws.cell(row=position, column=16).value = f'=$H{position}-$K{position}'
        self._ws.cell(row=position, column=17).value = f'=($H{position}-$K{position})/$H{position}'

        # 'Difference Good 1-4' formulas
        self._ws.cell(row=position, column=18).value = f'=$I{position}-$K{position}'
        self._ws.cell(row=position, column=19).value = f'=($I{position}-$K{position})/$I{position}'

        # 'Difference Redux to County' formulas
        self._ws.cell(row=position, column=20).value = f'=$L{position}-$N{position}'
        self._ws.cell(row=position, column=21).value = f'=($L{position}-$N{position})/$L{position}'

    def insert_summary_worksheet_rows(self, rows):
        self._rows = rows
        for i in range(len(self._rows)):
            self._insert_row(i+3, self._rows[i])

    def _set_header_styling(self):

        # header columns mapping
        header_columns = {
            'A1': '',

            # merged rows
            'B1': 'Link to G&B',
            'C1': 'Link to Submitted Case',
            'D1': 'Link to Single CMA Page',
            'E1': 'FOLIO',
            'F1': 'Petition Num',
            'G1': 'Tentative Value',
            'H1': 'All 1-4',
            'I1': 'Good 1-4',
            'J1': 'Redux Proposal',
            'K1': 'County Offer',
            'V1': 'Settle | Court',

            # merged columns
            'L1': 'Redux Discount',
            'N1': 'County Discount',
            'P1': 'Difference All 1-4',
            'R1': 'Difference Good 1-4',
            'T1': 'Difference Redux to County',
        }

        # not merged cells
        not_merged_cols = {
            'L2': '$',
            'M2': '%',
            'N2': '$',
            'O2': '%',
            'P2': '$',
            'Q2': '%',
            'R2': '$',
            'S2': '%',
            'T2': '$',
            'U2': '%'
        }

        # row cells merging
        self._ws.merge_cells('A1:A2')
        self._ws.merge_cells('B1:B2')
        self._ws.merge_cells('C1:C2')
        self._ws.merge_cells('D1:D2')
        self._ws.merge_cells('E1:E2')
        self._ws.merge_cells('F1:F2')
        self._ws.merge_cells('G1:G2')
        self._ws.merge_cells('H1:H2')
        self._ws.merge_cells('I1:I2')
        self._ws.merge_cells('J1:J2')
        self._ws.merge_cells('K1:K2')
        self._ws.merge_cells('V1:V2')

        # col cells merging
        self._ws.merge_cells('L1:M1')
        self._ws.merge_cells('N1:O1')
        self._ws.merge_cells('P1:Q1')
        self._ws.merge_cells('R1:S1')
        self._ws.merge_cells('T1:U1')

        # styling for merged header cells
        for key, value in header_columns.items():
            self._ws[key].style = 'Check Cell'
            self._ws[key].font = self.header_font
            self._ws[key].fill = self.header_fill
            self._ws[key].alignment = self.header_alignment
            self._ws[key].border = self.no_border
            self._ws[key].value = value

        # styling '%', '$' cols
        header_vertical_top_alignment = copy(self.header_alignment)
        header_vertical_top_alignment.vertical = 'top'

        for key, value in not_merged_cols.items():
            self._ws[key].style = 'Check Cell'
            self._ws[key].font = self.header_font
            self._ws[key].fill = self.header_fill
            self._ws[key].alignment = header_vertical_top_alignment
            self._ws[key].border = self.no_border
            self._ws[key].value = value

    def _as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def adjust_column_width(self, margin):
        for column_cells in self._ws.columns:
            length = max(len(self._as_text(cell.value)) for cell in column_cells)
            # self._ws.column_dimensions[column_cells[0].column].width = length + margin
            self._ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + margin

    def _set_columns_dimensions(self, margin=0.65):
        """
        Set columns dimensions
        """
        # self.adjust_column_width(margin)
        self._ws.column_dimensions['A'].width = 3.5 + margin
        self._ws.column_dimensions['B'].width = 15 + margin
        self._ws.column_dimensions['C'].width = 22 + margin
        self._ws.column_dimensions['D'].width = 22 + margin
        self._ws.column_dimensions['E'].width = 15 + margin
        self._ws.column_dimensions['F'].width = 15 + margin
        self._ws.column_dimensions['G'].width = 15 + margin
        self._ws.column_dimensions['H'].width = 15 + margin
        self._ws.column_dimensions['I'].width = 15 + margin
        self._ws.column_dimensions['J'].width = 15 + margin
        self._ws.column_dimensions['K'].width = 13 + margin
        self._ws.column_dimensions['L'].width = 15 + margin
        self._ws.column_dimensions['M'].width = 15 + margin
        self._ws.column_dimensions['N'].width = 15 + margin
        self._ws.column_dimensions['O'].width = 15 + margin
        self._ws.column_dimensions['P'].width = 15 + margin
        self._ws.column_dimensions['Q'].width = 15 + margin
        self._ws.column_dimensions['R'].width = 15 + margin
        self._ws.column_dimensions['S'].width = 15 + margin
        self._ws.column_dimensions['T'].width = 15 + margin
        self._ws.column_dimensions['U'].width = 15 + margin
        self._ws.column_dimensions['V'].width = 15 + margin

    def set_outline_border(self, cell_range):
        """
        Set outline border for cell range
        """
        rows = self._ws[cell_range]
        side = Side(border_style='medium', color="FF000000")
        rows = list(rows)

        # index of the last row
        max_y = len(rows) - 1
        for pos_y, cells in enumerate(rows):
            # index of the last cell
            max_x = len(cells) - 1
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    def _set_column_styling(self, max_row, min_col, max_col, min_row=3, style='Currency',
                            number_format=accounting_number_format, font=cell_font, alignment=cell_alignment,
                            border=None):
        """
        Set column(s):
            - style
            - number_format
            - font
            - alignment
            - border
        """
        for row in self._ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.style = style
                cell.number_format = number_format
                cell.font = font
                cell.alignment = alignment

                if border:
                    cell.border = border

    def _set_background_styling(self, row_count=10):
        """
        Set two colored rows background styling
        """
        for i in range(3, row_count + 1):
            if i % 2 == 0:
                for cell in self._ws[f"{i}:{i}"]:
                    cell.fill = self.background_fill


def send_devtools(driver, cmd, params={}):
    resource = "/session/%s/chromium/send_command_and_get_result" % driver.session_id
    url = driver.command_executor._url + resource
    body = json.dumps({'cmd': cmd, 'params': params})
    response = driver.command_executor._request('POST', url, body)
    if response.get('status'):
        raise Exception(response.get('value'))
    return response.get('value')


def save_as_pdf(driver, path, options={}):
    # https://timvdlippe.github.io/devtools-protocol/tot/Page#method-printToPDF
    result = send_devtools(driver, "Page.printToPDF", options)
    with open(path, 'wb') as file:
        file.write(base64.b64decode(result['data']))
